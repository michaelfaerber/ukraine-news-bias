import csv
import numpy as np
import openpyxl
import json
import copy
from numpy import mean
from collections import Counter

# THIS IS THE BASIC SCRIPT EXTRACTING ALL ANNOTATION FROM CROWD WORKERS
# ...AND PROVIDING ARTICLE OBJECTS FOR STATISTICS


path_full = r'..\reports\full'


jobs = ("subj", "hidden-assumpt", "framing", "bias")
# we splitted the annotation jobs in three groups according to the length of the articles to be labeled
groups = ("small", "medium", "long")
modi = ("avg", "maj", "intensified")


# Venezuela, Türkei, Mexiko, Vietnam, Indonesien, Ägypten, Bosnien und Herzegowina, Nordmazedonien, Argentinien, Peru, Kolumbien, Indien, Serbien, Montenegro
cw_rest = ["VEN", "TUR", "MEX", "VNM", "IDN", "EGY",
           "BIH", "MKD", "ARG", "PER", "COL", "IND", "SRB", "MNE"]

# United States, Rumänien, Großbritannien, Kanada, Italien, Spanien
cw_west = ["USA", "ROU", "GBR", "CAN", "ITA", "ESP"]

# Ukraine, Moldau
cw_sowj = ["UKR", "MDA"]

cw_groups = {
    "default": [],
    "west": cw_west,
    "sowj": cw_sowj,
    "rest": cw_rest
}

# used to extract judgements from reports received by Figure Eight (Appen)
label_postText = {
    "subj": "_please_rate_subjectivity_of_the_given_sentence",
    "hidden-assumpt": "_are_there_hidden_assumptions_present",
    "framing": {
        "russia": "_a_please_rate_the_representation_of_the_russian_government",
        "ukraine": "_b_please_rate_the_representation_of_the_ukrainian_government",
        "west": "_c_please_rate_the_representation_of_the_western_governments"
    },
    "bias": {
        "pro-russia": "_a_please_judge_whether_this_sentence_is_prorussia",
        "pro-west": "_b_please_judge_whether_this_sentence_is_prowest"
    }
}

# VALUE RANGES:
# Subj: 0 obj, 1 rather obj, 2 rather subj, 3 subj
# Hidden Assumpt: 0 no, 1 rather no, 2 rather yes, 3 yes
# Framing: 2 pos, 1 slightly pos, 0 neutral, -1 slightly neg, -2 neg
# Bias: 0 no, 1 rather no, 2 rather yes, 3 yes

# for each judgement, one country entry is entered in the array of the corresponding annotation job
crowdworkers = {
    "subj": [],
    "hidden-assumpt": [],
    "framing": [],
    "bias": []
}

articles = []

# CALL THIS FUNCTION by other scripts to receive article annotation.
# if there were no crowdworkers from the specified country group that judged sentences of an article,
# its article score is set to None.
def get_articles(cw_group):

    # articles and crowdworkers need to be cleared if method is called multiple times with different worker groups
    for key in crowdworkers.keys():
        crowdworkers[key] = []

    del articles[:]

    main(cw_group)

    return copy.deepcopy(articles), copy.deepcopy(crowdworkers)


def main(workers_group):

    get_all_annotation_in_objects(workers_group)
    calculate_sentence_scores(workers_group == "default")
    calculate_article_scores()
    get_expert_leaning_per_article()
    write_json_files()
    # write_sent_annotation_in_file()
    # write_labels_in_file()
    print("Created all article objects!")

    return articles, crowdworkers


def count_sentences(articles):
    # call this function if you want to receive the number of all sentences

    sent_counter = sum([len(a.sentences) for a in articles])
    print("Counted Sentences: " + str(sent_counter))


def get_all_annotation_in_objects(group_name):

    def workerInGroup(country):
        if group_name == "default":
            return True
        else:
            return country in cw_groups[group_name]

    for job in jobs:
        for group in groups:
            this_path = path_full + "\\" + job + "-" + group + ".csv"
            file = open(this_path, "r", encoding='utf-8')
            csv_reader = csv.DictReader(file, delimiter=",")

            for row in csv_reader:
                # one article too much for pro-Russia
                article, toAdd = get_article_ref(row)
                rightWorker = workerInGroup(row['_country'])

                if toAdd:
                    if rightWorker:  # filter out judgements from workers of uninterested countries
                        worker_id = row['_worker_id']
                        worker = next(
                            (x for x in crowdworkers[job] if x.id == worker_id), None)
                        if worker is not None:
                            worker.increment_judgements()
                        else:
                            worker = Crowdworker(row['_country'], worker_id)
                            crowdworkers[job].append(worker)

                    # add sentences but not rating if worker from uninterested countries
                    get_sentences_in_art_object(
                        job, group, article, row, rightWorker, row['_country'])

            file.close()

    print(f'Created article objects: {len(articles)} current articles.')


def get_sentences_in_art_object(job, group, article, row, addRating, origin):
    # for the framing annotation of long news articles, we had to split each article in eigh parts
    # ...resulting in a more complicated annotation extraction
    if (group == "long") and (job == "framing"):
        parts = {
            'ONE': range(0, 11),  # includes title, sent_0 to sent_9
            'TWO': range(11, 21),  # includes only sent_0 to sent_9
            'THREE': range(21, 31),
            'FOUR': range(31, 41),
            'FIVE': range(41, 51),
            'SIX': range(51, 61),
            'SEVEN': range(61, 71),
            'EIGHT': range(71, 81),
        }

        for i in range(81):

            my_range = parts[row['part']]

            if i in my_range:
                number = (9, i % 10 - 1)[i % 10 - 1 > -1]
                index = ('sent_' + str(number), "title")[i == 0]
                position = i

                if row[index] != '':
                    sent = get_sentence_ref(row[index], article, position)
                    if row['part'] != 'ONE':  # add 1 for title
                        number = number + 1
                    if addRating:
                        sent.add_rating(job, number, row, origin)

    else:
        for i in range(81):
            index = ("sent_" + str(i-1), "title")[i == 0]
            if article.article_id == 'ukr_1932' and i == 52:  # wrong content with 'contact the author...'
                break
            if row[index] != '':
                sent = get_sentence_ref(row[index], article, i)
                if addRating:
                    sent.add_rating(job, i, row, origin)
            else:
                break


def write_json_files():
    # this function generates the published data set

    sentence_counter = 0

    for i, article in enumerate(articles):
        data = {}
        data['id'] = "article_" + str(i)
        data['sentences'] = []

        framing = {'russia': {}, 'ukraine': {}, 'west': {}}
        bias = {'pro-russia': {}, 'pro-west': {}}

        for gov in framing:
            framing[gov]['pos_sent'] = {}
            framing[gov]['neutral_sent'] = {}
            framing[gov]['neg_sent'] = {}

        for mode in modi:
            for gov in framing:
                framing[gov]['pos_sent'][mode] = article.scores['framing'][gov][mode][0][1]
                framing[gov]['neutral_sent'][mode] = article.scores['framing'][gov][mode][1][1]
                framing[gov]['neg_sent'][mode] = article.scores['framing'][gov][mode][2][1]

            for key in bias:
                bias[key][mode] = article.scores['bias'][key][mode]

        data['subjectivity'] = article.scores["subj"]
        data['hidden_assumptions'] = article.scores["hidden-assumpt"]
        data['framing'] = framing

        def get_sent_obj(sent, art_id):
            current_sent = {}
            current_sent['article'] = art_id
            current_sent['id'] = 'sent_' + str(sentence_counter)
            current_sent['position_in_article'] = sent.position
            current_sent['content'] = sent.content
            current_sent['subjectivity'] = {
                "score": sent.scores['subj'],
                "judgements": sent.values['subj'],
                "cw_origins": sent.cw_origins['subj']
            }
            current_sent['hidden_assumptions'] = {
                "score": sent.scores['hidden-assumpt'],
                "judgements": sent.values['hidden-assumpt'],
                "cw_origins": sent.cw_origins['hidden-assumpt']
            }
            current_sent['framing'] = {
                "score": sent.scores['framing'],
                "judgements": sent.values['framing'],
                "cw_origins": sent.cw_origins['framing']
            }
            current_sent['bias'] = {
                "score": sent.scores['bias'],
                "judgements": sent.values['bias'],
                "cw_origins": sent.cw_origins['bias']
            }

            return current_sent

        for mySent in article.sentences:
            data['sentences'].append(get_sent_obj(mySent, data['id']))
            sentence_counter += 1

        with open('all-data-as-json/' + data['id'] + '.json', 'w') as outfile:
            json.dump(data, outfile)


def write_sent_annotation_in_file():
    # deprecated (regards only bias dimensions, not bias itself and outputs csv file)

    file_path = "full-sentence-annotation.csv"
    with open(file_path, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(
            ["article_id", "sent_position", "sent_content",
             "subj_contributor_1", "subj_contributor_2", "subj_contributor_3", "subj_contributor_4", "subj_contributor_5",
             "hidden_assumpt_contributor_1", "hidden_assumpt_contributor_2", "hidden_assumpt_contributor_3", "hidden_assumpt_contributor_4", "hidden_assumpt_contributor_5",
             "framing_russia_contributor_1", "framing_russia_contributor_2", "framing_russia_contributor_3", "framing_russia_contributor_4", "framing_russia_contributor_5",
             "framing_ukraine_contributor_1", "framing_ukraine_contributor_2", "framing_ukraine_contributor_3", "framing_ukraine_contributor_4", "framing_ukraine_contributor_5",
             "framing_west_contributor_1", "framing_west_contributor_2", "framing_west_contributor_3", "framing_west_contributor_4", "framing_west_contributor_5"
             ]
        )
        for art in articles:
            for sent in art.sentences:
                current_row = [art.article_id, sent.position, sent.content]

                for val_subj in sent.subj_val:
                    current_row.append(val_subj)

                for val_hid in sent.hidden_assumpt_val:
                    current_row.append(val_hid)

                for key in sent.framing_val:
                    current_row += sent.framing_val[key]

                writer.writerow(current_row)


def write_labels_in_file():
    # deprecated (regards only bias dimensions, not bias itself and outputs csv file)

    file_path = "article-labels.csv"
    with open(file_path, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(
            ["article_id", "subj_avg", "subj_maj", "subj_intensified", "hidden_assumpt_avg", "hidden_assumpt_maj", "hidden_assumpt_intensified",
             "framing_russia_avg", "framing_russia_maj", "framing_russia_intensified",
             "framing_ukraine_avg", "framing_ukraine_maj", "framing_ukraine_intensified",
             "framing_west_avg", "framing_west_maj", "framing_west_intensified"
             ]
        )

        for art in articles:
            current_row = [art.article_id]
            all_mode_scores = [art.scores[job][score]
                               for job in jobs for score in art.scores[job]]
            for i, mode in enumerate(modi):
                all_mode_scores[i+6] = all_mode_scores[i+6][mode]
            current_row += all_mode_scores
            writer.writerow(current_row)

    file_path = "sentence-labels.csv"
    with open(file_path, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(
            ["article_id", "sent_id", "subj_avg", "subj_maj", "subj_intensified", "hidden_assumpt_avg", "hidden_assumpt_maj", "hidden_assumpt_intensified",
             "framing_russia_avg", "framing_russia_maj", "framing_russia_intensified",
             "framing_ukraine_avg", "framing_ukraine_maj", "framing_ukraine_intensified",
             "framing_west_avg", "framing_west_maj", "framing_west_intensified"
             ]
        )

        for art in articles:
            for sent in art.sentences:
                current_row = [art.article_id,
                               art.article_id + "_" + str(sent.position)]
                all_mode_scores = [sent.scores[job][score]
                                   for job in jobs for score in sent.scores[job]]

                for i, mode in enumerate(modi):
                    all_mode_scores[i+6] = all_mode_scores[i+6][mode]

                current_row += all_mode_scores
                writer.writerow(current_row)


def calculate_sentence_scores(all_modes):

    # helper method, use when necessary
    def check_if_empty(vals):
        if (vals == []):
            print("Sent: " + str(sent.position) +
                  ", Article: " + article.article_id)

    for job in jobs:
        for article in articles:
            for sent in article.sentences:
                if not all_modes:
                    modi = ["avg", "maj"]
                else:
                    modi = ["avg", "maj", "intensified"]

                for mode in modi:
                    if job == "framing" or job == "bias":
                        for key in sent.values[job]:
                            current_values = sent.values[job][key]
                            current_scores = sent.scores[job][key]
                            # check_if_empty(current_values)
                            insert_score_helper(
                                current_values, current_scores, mode, job)
                    else:
                        current_values = sent.values[job]
                        # check_if_empty(current_values)
                        current_scores = sent.scores[job]
                        insert_score_helper(
                            current_values, current_scores, mode, job)
            print("Added all scores to article " + article.article_id)


def calculate_article_scores():

    for article in articles:
        scores_obj = Sent_Score_Obj()

        for mode in modi:
            for sent in article.sentences:
                for job in jobs:
                    if job == "framing" or job == "bias":
                        for key in sent.scores[job]:
                            sent_scores = sent.scores[job][key]
                            scores_obj.all_scores[job][key][mode].append(
                                sent_scores[mode])
                    else:
                        sent_scores = sent.scores[job]
                        scores_obj.all_scores[job][mode].append(
                            sent_scores[mode])
        insert_article_score_helper(article, scores_obj.all_scores)
        print("Calculated scores for article " + article.article_id)


def get_expert_leaning_per_article():

    def get_leaning_helper(article_id, rows):
        for row in rows:
            if row[0].value == article_id:
                return int(row[8].value)
        print("Article Leaning not found!")

    # you need access to the original data set of Cremisini et al. in order to receive the expert article leaning
    wb = openpyxl.load_workbook(
        filename=r'..\..\..\Ukraine-Crisis-Dataset\mediabias-dataset-andres\ukr_final_DAG.xlsx')
    sheet = wb.active

    for article in articles:
        current_id = article.article_id
        article.leaning = get_leaning_helper(current_id, sheet.iter_rows(
            max_col=sheet.max_column))  # new generator each time

        # print("Added leaning of " + str(article.leaning) +
        #       " to " + article.article_id)

    for article in articles:
        if article.leaning not in [0, 1, -1]:
            print(article.article_id)


def insert_score_helper(value_list, score_list, current_mode, current_job):
    # attention, the score is set to None if no judgements for this sentence is given

    if current_mode == "avg":
        mean_score = (None, mean(value_list))[len(value_list) != 0]
        score_list["avg"] = mean_score

    else:

        if current_mode == "maj":
            score_list["maj"] = perform_maj_vote(value_list)

        else:
            # current_mode: "intensified"

            # check if two crowdworkers selected non-neutral answer
            less_zero_list = list(filter(lambda x: x < 0, value_list))
            greater_one_list = list(filter(lambda x: x > 1, value_list))
            greater_zero_list = list(filter(lambda x: x > 0, value_list))

            longest_list = max(
                [less_zero_list, greater_zero_list], key=lambda l: len(l))

            conditions = {
                "subj": (len(greater_one_list) > 1, greater_one_list),
                "hidden-assumpt": (len(greater_one_list) > 1, greater_one_list),
                "framing": (True, longest_list) if (len(longest_list) > 1) else (False, []),
                "bias": (len(greater_one_list) > 1, greater_one_list)
            }
            true_condition, my_values = conditions[current_job]

            intensified_value = mean(
                my_values) if true_condition else perform_maj_vote(value_list)

            score_list["intensified"] = intensified_value


def insert_article_score_helper(current_article, all_scores):

    for mode in modi:
        for job in jobs:

            if job == "framing" or job == "bias":
                for key in all_scores[job]:
                    my_scores = []
                    score_list = all_scores[job][key][mode]

                    if None in score_list:
                        current_article.scores[job][key][mode] = None
                    else:
                        if job == "framing":
                            target_groups = {
                                "pos_sent": lambda x: x > 0,
                                "neutral_sent": lambda x: x == 0,
                                "neg_sent": lambda x: x < 0
                            }
                            for groupname, condition in target_groups.items():
                                sent_group = list(
                                    filter(condition, score_list))
                                group_score = len(sent_group) / len(score_list)
                                my_scores.append((groupname, group_score))
                        else:
                            biased_sent = list(
                                filter(lambda x: x > 1, score_list))
                            my_scores = len(biased_sent) / len(score_list)

                        current_article.scores[job][key][mode] = my_scores

            else:
                score_list = all_scores[job][mode]

                if None in score_list:
                    current_article.scores[job][mode] = None
                else:
                    # calculate share of "biased sentences"
                    biased_sent = list(filter(lambda x: x > 1, score_list))
                    my_score = len(biased_sent) / len(score_list)

                    current_article.scores[job][mode] = my_score


def perform_maj_vote(my_value_list):

    if len(my_value_list) > 1:
        counts = Counter(my_value_list)
        majorities = counts.most_common(2)

        if len(majorities) == 1:
            value = majorities[0][0]
        # if all contributors selected different answers (only possible with framing job):
        elif majorities[0][1] == 1:
            value = 0.0  # neutral
        elif majorities[0][1] == majorities[1][1]:
            value = mean([majorities[0][0], majorities[1][0]])
        else:
            value = majorities[0][0]

        return value

    elif len(my_value_list) == 1:
        return my_value_list[0]

    else:
        return None


def get_article_ref(data_row):

    art_id = data_row['article_id']

    if art_id == 'ukr_1960':  # exclude this article
        return None, False

    # find corresponding article in article list
    current_art = next((a for a in articles if a.article_id == art_id), None)
    art_found = current_art is not None

    # create new article if not already there and keep reference
    if not art_found:
        current_art = Article(data_row['article_id'])
        articles.append(current_art)

    return current_art, True


def get_sentence_ref(text, current_art, position):

    # find corresponding sentence of article
    current_sent = next(
        (sent for sent in current_art.sentences if sent.position == position), None)
    sent_found = current_sent is not None

    # otherwise create new sentence obj
    if not sent_found:
        current_sent = Sentence(position, text)
        current_art.sentences.append(current_sent)

    return current_sent


class Sent_Score_Obj:
    # helper class

    def __init__(self):
        self.all_scores = {
            "subj": {
                "avg": [],
                "maj": [],
                "intensified": []
            },
            "hidden-assumpt": {
                "avg": [],
                "maj": [],
                "intensified": []
            },
            "bias": {
                "pro-russia": {"avg": [],
                               "maj": [],
                               "intensified": []},
                "pro-west": {"avg": [],
                             "maj": [],
                             "intensified": []},
            },
            "framing": {
                "russia": {"avg": [],
                           "maj": [],
                           "intensified": []},
                "ukraine": {"avg": [],
                            "maj": [],
                            "intensified": []},
                "west": {"avg": [],
                         "maj": [],
                         "intensified": []}
            }
        }


class Sentence:

    def __init__(self, pos=0, text=""):
        self.content = text
        self.position = pos
        self.subj_val = []
        self.hidden_assumpt_val = []
        self.bias_val = {
            "pro-russia": [],
            "pro-west": []
        }
        self.framing_val = {
            "russia": [],
            "ukraine": [],
            "west": []
        }

        self.values = {
            "subj": self.subj_val,
            "hidden-assumpt": self.hidden_assumpt_val,
            "bias": self.bias_val,
            "framing": self.framing_val
        }

        self.cw_origins = {
            "subj": [],
            "hidden-assumpt": [],
            "bias": [],
            "framing": []
        }

        self.scores = {
            "subj": {
                "avg": "",
                "maj": "",
                "intensified": ""
            },
            "hidden-assumpt": {
                "avg": "",
                "maj": "",
                "intensified": ""
            },
            "bias": {
                "pro-russia": {
                    "avg": "",
                    "maj": "",
                    "intensified": ""},
                "pro-west": {
                    "avg": "",
                    "maj": "",
                    "intensified": ""
                }
            },
            "framing": {
                "russia": {"avg": "",
                           "maj": "",
                           "intensified": ""},
                "ukraine": {"avg": "",
                            "maj": "",
                            "intensified": ""},
                "west": {"avg": "",
                         "maj": "",
                         "intensified": ""}
            }
        }

    def add_rating(self, job, index, data_row, origin):

        currentList = self.values[job]
        if job == "hidden-assumpt" or job == "subj":
            value = float(data_row[str(index+1) + label_postText[job]])

            if job == "subj":
                change_Val = {  # subj has different range values from figure 8 (appen)
                    2.0: 0.0,
                    1.0: 1.0,
                    -1.0: 2.0,
                    -2.0: 3.0
                }
                value = change_Val[value]

            currentList.append(value)
        else:
            # job == "framing" or job == "bias"
            for key in currentList:
                value = float(
                    data_row[str(index+1) + label_postText[job][key]])
                currentList[key].append(value)

        self.cw_origins[job].append(origin)

    def is_biased(self, job, mode):
        # call this function in scripts to determine if a sentence is labeled as biased
        # ... w.r.t. bias dimension or bias itself

        biased = False
        for key in self.scores[job]:
            if key in modi:
                # if subjectiv, then score > 1
                # if hidden assumpt, then score > 1
                biased = self.scores[job][mode] > 1
                break
            else:
                if job == "bias":
                    # if bias, then at least one tendency score has to be not neutral
                    biased = biased or (self.scores[job][key][mode] > 1)
                else:
                    # if framing, then at least one government has to be represented not neutral
                    biased = biased or (self.scores[job][key][mode] != 0)

        return biased


class Article:

    def __init__(self, id="test_1"):
        self.article_id = id
        self.sentences = []
        self.leaning = 10
        self.scores = {
            "subj": {
                "avg": "",
                "maj": "",
                "intensified": ""
            },
            "hidden-assumpt": {
                "avg": "",
                "maj": "",
                "intensified": ""
            },
            "bias": {
                "pro-russia": {
                    "avg": "",
                    "maj": "",
                    "intensified": ""},
                "pro-west": {
                    "avg": "",
                    "maj": "",
                    "intensified": ""
                }
            },
            "framing": {
                "russia": {"avg": [],
                           "maj": [],
                           "intensified": []},
                "ukraine": {"avg": [],
                            "maj": [],
                            "intensified": []},
                "west": {"avg": [],
                         "maj": [],
                         "intensified": []}
            }
        }


class Crowdworker:
    judgements = 0

    def __init__(self, country, id):
        self.id = id
        self.country = country

    def increment_judgements(self):
        self.judgements += 1


# IF YOU ONLY WANT TO RUN THIS SCRIPT, UNCOMMENT THE FOLLOWING
# get_articles("default")
